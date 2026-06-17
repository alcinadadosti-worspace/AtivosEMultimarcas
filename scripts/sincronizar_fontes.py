# -*- coding: utf-8 -*-
"""
Sincroniza as planilhas-fonte versionadas com o estado corrigido de
data/produtos.db, e PROVA a consistência reconstruindo um banco do zero.

1. estoqueplanilha.xlsx: atualiza a Marca dos SKUs cuja marca corrigida
   (em produtos) difere — mesmo após normalizar — da que está na planilha.
   (Captura as 18 correções reais; ignora apelidos que o import já normaliza.)
2. make_iaf.xlsx: acrescenta as maquiagens que estão em iaf_make e faltam no xlsx.
3. Verificação: importa as planilhas para um banco temporário e compara.
"""
import os
import sys
import shutil
import sqlite3
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from app.utils.normalizers import normalizar_sku, normalizar_marca
import import_db

DB = os.path.join("data", "produtos.db")
ESTOQUE = os.path.join("data", "estoqueplanilha.xlsx")
MAKE_IAF = os.path.join("data", "make_iaf.xlsx")


def estado_db():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("SELECT sku_normalizado, sku, nome, marca FROM produtos")
    prod = {r[0]: (r[1], r[2], r[3]) for r in cur.fetchall()}
    cur.execute("SELECT sku_normalizado, sku, descricao, marca FROM iaf_make")
    iafm = {r[0]: (r[1], r[2], r[3]) for r in cur.fetchall()}
    conn.close()
    return prod, iafm


def sync_estoque(prod):
    wb = openpyxl.load_workbook(ESTOQUE)
    ws = wb.active
    hdr = [str(c.value or "").strip().lower() for c in ws[1]]
    c_sku = hdr.index("sku") + 1
    c_marca = hdr.index("marca") + 1
    alterados = 0
    for ri in range(2, ws.max_row + 1):
        skn = normalizar_sku(ws.cell(ri, c_sku).value)
        if not skn or skn not in prod:
            continue
        marca_correta = prod[skn][2]
        marca_atual = ws.cell(ri, c_marca).value
        # só altera quando é correção real (não mero apelido que o import normaliza)
        if normalizar_marca(marca_atual) != marca_correta:
            ws.cell(ri, c_marca, value=marca_correta)
            alterados += 1
    wb.save(ESTOQUE)
    return alterados


def append_make_iaf(iafm):
    wb = openpyxl.load_workbook(MAKE_IAF)
    ws = wb.active
    hdr = [str(c.value or "").strip().lower() for c in ws[1]]
    c_cod = 1  # Codigo
    existentes = set()
    for ri in range(2, ws.max_row + 1):
        existentes.add(normalizar_sku(ws.cell(ri, c_cod).value))
    novos = 0
    for skn, (sku, desc, marca) in iafm.items():
        if skn in existentes:
            continue
        ws.append([sku, desc, marca])
        novos += 1
    wb.save(MAKE_IAF)
    return novos


def verificar(prod, iafm):
    """Reconstrói um banco do zero a partir das planilhas e compara."""
    tmp = tempfile.mktemp(suffix=".db")
    conn = sqlite3.connect(tmp)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE produtos (id INTEGER PRIMARY KEY AUTOINCREMENT, sku TEXT, sku_normalizado TEXT UNIQUE, nome TEXT, marca TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cur.execute("""CREATE TABLE iaf_make (id INTEGER PRIMARY KEY AUTOINCREMENT, sku TEXT, sku_normalizado TEXT, descricao TEXT, marca TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    conn.commit()
    from pathlib import Path
    import_db.import_produtos(conn, Path(ESTOQUE))
    import_db.import_iaf(conn, Path(MAKE_IAF), "iaf_make")

    cur.execute("SELECT marca, COUNT(*) FROM produtos GROUP BY marca ORDER BY 2 DESC")
    dist_novo = dict(cur.fetchall())
    cur.execute("SELECT COUNT(*) FROM iaf_make")
    iaf_novo = cur.fetchone()[0]

    # comparar as 18 correções: cada SKU corrigido bate?
    divergencias = []
    for skn, (sku, nome, marca) in prod.items():
        cur.execute("SELECT marca FROM produtos WHERE sku_normalizado=?", (skn,))
        r = cur.fetchone()
        if r and r[0] != marca:
            divergencias.append((skn, marca, r[0]))
    conn.close()
    os.remove(tmp)
    return dist_novo, iaf_novo, divergencias


def main():
    shutil.copy2(MAKE_IAF, MAKE_IAF + ".bak_pre_sync")
    print(f"[backup] {MAKE_IAF}.bak_pre_sync (estoque já tem backup do passo anterior)")

    prod, iafm = estado_db()
    n_est = sync_estoque(prod)
    print(f"[estoqueplanilha] marcas corrigidas: {n_est}")
    n_iaf = append_make_iaf(iafm)
    print(f"[make_iaf] maquiagens acrescentadas: {n_iaf}")

    print("\n=== VERIFICAÇÃO: reconstruindo banco do zero a partir das planilhas ===")
    dist, iaf_novo, div = verificar(prod, iafm)
    print("  Distribuição por marca (reconstruído):")
    for m, q in dist.items():
        print(f"    {m}: {q}")
    print(f"  iaf_make reconstruído: {iaf_novo}")
    print(f"  Divergências de marca vs produtos.db corrigido: {len(div)}")
    for skn, esperado, obtido in div[:20]:
        print(f"    {skn}: esperado {esperado!r} | obtido {obtido!r}")


if __name__ == "__main__":
    main()

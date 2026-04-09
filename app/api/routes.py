"""
FastAPI routes for Multimarks Analytics.
"""
import base64
import json as _json
import os
import sqlite3
import threading
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Cookie, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import JSONResponse, Response

import polars as pl
from openpyxl import Workbook, load_workbook

from app.api.dependencies import get_db
from app.config import DATA_DIR, GEO_PARQUET_PATH, GEO_STATS_PATH, MARCAS_GRUPO, MOTIVO_MATCH_EXATO
from app.utils.normalizers import normalizar_sku
from app.api.schemas import (
    UploadResponse,
    MetricasGerais,
    FiltrosDisponiveis,
    HealthCheck,
)
from app.services.session import (
    get_session,
    get_session_data as get_session_by_id,
    set_session_value,
    clear_session,
    get_session_stats,
)
from app.services.venda import (
    processar_planilha_vendas,
    obter_ciclos_unicos,
    obter_setores_unicos,
    obter_marcas_unicas,
    obter_gerencias_unicas,
)
from app.services.metricas import (
    calcular_metricas_cliente,
    calcular_metricas_gerais,
    calcular_vendas_por_marca,
    calcular_top_setores,
    calcular_top_setores_completo,
    calcular_evolucao_ciclos,
    calcular_resumo_ciclos,
    calcular_dados_setor_ciclo,
    calcular_combinacoes_marcas,
    calcular_metricas_por_setor,
    aplicar_filtros,
    obter_detalhes_cliente,
)
from app.services.iaf import (
    cruzar_vendas_com_iaf,
    calcular_percentual_iaf,
    calcular_iaf_por_setor,
    listar_vendas_iaf,
)
from app.services.auditoria import (
    obter_estatisticas_auditoria,
    listar_auditoria,
    listar_produtos_novos,
)
from app.services.metas import ler_planilha_metas, encontrar_meta_setor
from app.services.categoria import (
    classificar_vendas,
    calcular_metricas_categoria,
    calcular_categoria_por_ciclo,
    calcular_categoria_por_setor,
    listar_produtos_categoria,
    obter_categorias_disponiveis,
)
from app.services.ranking import (
    calcular_ranking_revendedoras,
    calcular_evolucao_revendedora,
    calcular_comparativo_ciclos,
)
from app.utils.exporters import exportar_csv, exportar_excel, exportar_multiplas_abas
from app.services.geo import (
    processar_planilha_clientes,
    calcular_metricas_bairro,
    calcular_detalhe_bairro,
    calcular_metricas_cidade,
    listar_clientes_geo,
    obter_cidades_geo,
    obter_bairros_geo,
)


# Router for API endpoints
api_router = APIRouter(prefix="/api")

# Cookie configuration
SESSION_COOKIE_NAME = "multimarks_session"
SESSION_COOKIE_MAX_AGE = 60 * 60 * 24  # 24 hours
ESTOQUE_PLANILHA_PATH = DATA_DIR / "estoqueplanilha.xlsx"


def _find_header_indexes(header_row: List[Any]) -> Dict[str, int]:
    """Find SKU, name and brand columns in worksheet header."""
    indexes: Dict[str, int] = {}
    for idx, value in enumerate(header_row, start=1):
        col = str(value or "").strip().lower()
        if col in ("sku", "codigo", "código", "cod", "codigoproduto") and "sku" not in indexes:
            indexes["sku"] = idx
        elif col in ("nome", "nomeproduto", "descricao", "descrição", "produto") and "nome" not in indexes:
            indexes["nome"] = idx
        elif col == "marca" and "marca" not in indexes:
            indexes["marca"] = idx
    return indexes


def _upsert_produtos_na_planilha(produtos: List[Dict[str, str]]) -> None:
    """
    Persist products in data/estoqueplanilha.xlsx.

    The spreadsheet is treated as source-of-truth backup so manually saved
    products are not lost if the SQLite DB is rebuilt later.
    """
    path: Path = ESTOQUE_PLANILHA_PATH
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        header_values = [cell.value for cell in ws[1]]
        header_idx = _find_header_indexes(header_values)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "produtos"
        ws.append(["sku", "nome", "marca"])
        header_idx = {"sku": 1, "nome": 2, "marca": 3}

    if not {"sku", "nome", "marca"}.issubset(header_idx):
        raise ValueError("Planilha de estoque sem cabecalho esperado (sku, nome, marca)")

    sku_col = header_idx["sku"]
    nome_col = header_idx["nome"]
    marca_col = header_idx["marca"]

    # Build index of existing normalized SKUs in spreadsheet.
    rows_by_sku_norm: Dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        sku_value = ws.cell(row=row_idx, column=sku_col).value
        sku_norm = normalizar_sku(sku_value)
        if sku_norm:
            rows_by_sku_norm[sku_norm] = row_idx

    for produto in produtos:
        sku = produto["sku"]
        nome = produto["nome"]
        marca = produto["marca"]
        sku_norm = normalizar_sku(sku)
        if not sku_norm:
            continue

        row_idx = rows_by_sku_norm.get(sku_norm)
        if row_idx is None:
            row_idx = ws.max_row + 1
            rows_by_sku_norm[sku_norm] = row_idx

        ws.cell(row=row_idx, column=sku_col, value=sku)
        ws.cell(row=row_idx, column=nome_col, value=nome.strip())
        ws.cell(row=row_idx, column=marca_col, value=marca)

    wb.save(path)


def _backup_para_github() -> None:
    """
    Push estoqueplanilha.xlsx to GitHub via Contents API so it survives Render deploys.

    Requires env vars:
      GITHUB_TOKEN  — personal access token with repo write access
      GITHUB_REPO   — owner/repo (e.g. "usuario/meu-repo")
      GITHUB_BRANCH — branch to commit to (default: "main")
      GITHUB_ESTOQUE_PATH — path of the file inside the repo (default: "data/estoqueplanilha.xlsx")

    Silent no-op when env vars are missing or on any error.
    """
    token = os.environ.get("GITHUB_TOKEN", "")
    repo = os.environ.get("GITHUB_REPO", "")
    branch = os.environ.get("GITHUB_BRANCH", "main")
    file_path_in_repo = os.environ.get("GITHUB_ESTOQUE_PATH", "data/estoqueplanilha.xlsx")

    if not token or not repo:
        return

    path = ESTOQUE_PLANILHA_PATH
    if not path.exists():
        return

    try:
        with open(path, "rb") as f:
            content_b64 = base64.b64encode(f.read()).decode("utf-8")

        api_url = f"https://api.github.com/repos/{repo}/contents/{file_path_in_repo}"
        headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github+json",
            "Content-Type": "application/json",
        }

        # Get current SHA (required for updates)
        sha = None
        try:
            req = urllib.request.Request(api_url, headers=headers)
            with urllib.request.urlopen(req) as resp:
                sha = _json.loads(resp.read()).get("sha")
        except urllib.error.HTTPError as e:
            if e.code != 404:
                return  # unexpected error, abort

        body: Dict[str, str] = {
            "message": "backup: atualizar estoqueplanilha.xlsx",
            "content": content_b64,
            "branch": branch,
        }
        if sha:
            body["sha"] = sha

        req = urllib.request.Request(
            api_url,
            data=_json.dumps(body).encode("utf-8"),
            headers=headers,
            method="PUT",
        )
        urllib.request.urlopen(req)
    except Exception:
        pass  # fire-and-forget — never break the user flow


def _atualizar_sessao_com_produtos_cadastrados(
    session_id: str,
    session_data: Dict[str, Any],
    produtos: List[Dict[str, str]],
) -> int:
    """
    Update current session data so newly registered products are recognized immediately.

    Returns:
        Number of rows updated in df_vendas.
    """
    df_vendas = session_data.get("df_vendas")
    if df_vendas is None or df_vendas.is_empty() or not produtos:
        return 0

    sku_para_marca: Dict[str, str] = {}
    sku_para_nome: Dict[str, str] = {}
    for p in produtos:
        sku_norm = normalizar_sku(p.get("sku"))
        if not sku_norm:
            continue
        sku_para_marca[sku_norm] = p.get("marca", "")
        sku_para_nome[sku_norm] = p.get("nome", "")

    if not sku_para_marca:
        return 0

    skus_norm = list(sku_para_marca.keys())
    mapa_marca = pl.DataFrame({
        "CodigoProduto_normalizado": skus_norm,
        "_marca_nova": [sku_para_marca[s] for s in skus_norm],
        "_nome_novo": [sku_para_nome[s] for s in skus_norm],
    })

    df_join = df_vendas.join(mapa_marca, on="CodigoProduto_normalizado", how="left")
    df_vendas_atualizado = df_join.with_columns([
        pl.when(pl.col("_marca_nova").is_not_null())
          .then(pl.col("_marca_nova"))
          .otherwise(pl.col("Marca_BD"))
          .alias("Marca_BD"),
        pl.when(pl.col("_nome_novo").is_not_null())
          .then(pl.col("_nome_novo"))
          .otherwise(pl.col("Nome_BD"))
          .alias("Nome_BD"),
        pl.when(pl.col("_marca_nova").is_not_null())
          .then(pl.lit(MOTIVO_MATCH_EXATO))
          .otherwise(pl.col("Motivo_Match"))
          .alias("Motivo_Match"),
    ]).drop(["_marca_nova", "_nome_novo"])

    linhas_atualizadas = len(
        df_vendas.filter(pl.col("CodigoProduto_normalizado").is_in(skus_norm))
    )

    # Keep session metrics consistent with the updated brand matches.
    df_clientes_atualizado = calcular_metricas_cliente(df_vendas_atualizado)
    set_session_value(session_id, "df_vendas", df_vendas_atualizado)
    set_session_value(session_id, "df_clientes", df_clientes_atualizado)

    return linhas_atualizadas


def get_user_session(session_id: Optional[str] = Cookie(None, alias=SESSION_COOKIE_NAME)):
    """
    Dependency to get the current user's session.

    Returns tuple of (session_id, session_data).
    Creates new session if none exists.
    """
    return get_session(session_id)


def create_response_with_session(data: Any, session_id: str) -> JSONResponse:
    """Create a JSON response with session cookie set."""
    response = JSONResponse(content=data)
    response.set_cookie(
        key=SESSION_COOKIE_NAME,
        value=session_id,
        max_age=SESSION_COOKIE_MAX_AGE,
        httponly=True,
        samesite="lax",
    )
    return response


# Legacy function for compatibility with page routes
def get_session_data():
    """
    Legacy function - returns empty dict for page route compatibility.
    Pages should check has_data via API instead.
    """
    return {"df_vendas": None}


# =============================================================================
# HEALTH CHECK
# =============================================================================

@api_router.api_route("/health", methods=["GET", "HEAD"], response_model=HealthCheck)
async def health_check():
    """Health check endpoint for Uptime Robot."""
    return {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat(),
        "service": "multimarks-analytics"
    }


@api_router.post("/clear")
async def clear_cache(
    session: tuple = Depends(get_user_session),
):
    """Clear session cache and force reload on next upload."""
    session_id, _ = session
    clear_session(session_id)
    return create_response_with_session({
        "success": True,
        "message": "Cache limpo. Faca upload novamente da planilha."
    }, session_id)


@api_router.get("/clear")
async def clear_cache_get(
    session: tuple = Depends(get_user_session),
):
    """Clear session cache (GET method for easy browser access)."""
    session_id, _ = session
    clear_session(session_id)
    return create_response_with_session({
        "success": True,
        "message": "Cache limpo. Faca upload novamente da planilha."
    }, session_id)


@api_router.get("/session/stats")
async def session_stats():
    """Get session statistics (for monitoring)."""
    return get_session_stats()


@api_router.get("/session/status")
async def session_status(
    session: tuple = Depends(get_user_session),
):
    """Check if current session has data loaded."""
    session_id, session_data = session
    has_data = session_data.get("df_vendas") is not None

    response = create_response_with_session({
        "has_data": has_data,
        "session_id": session_id[:8] + "...",  # Only show first 8 chars for privacy
    }, session_id)

    return response


# =============================================================================
# UPLOAD
# =============================================================================

@api_router.post("/upload")
async def upload_vendas(
    file: UploadFile = File(...),
    conn: sqlite3.Connection = Depends(get_db),
    session: tuple = Depends(get_user_session),
):
    """
    Upload and process a sales spreadsheet.

    Accepts CSV or Excel files with required columns.
    Automatically clears previous session data before processing.
    """
    session_id, session_data = session

    # Clear previous session data automatically
    clear_session(session_id)

    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo nao fornecido")

    # Validate file extension
    if not file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Formato de arquivo invalido. Use CSV ou Excel (.xlsx, .xls)"
        )

    try:
        # Read file content
        content = await file.read()

        # Process spreadsheet
        resultado = processar_planilha_vendas(content, file.filename, conn)

        # Calculate customer metrics
        df_clientes = calcular_metricas_cliente(resultado["df_vendas"])

        # Store in user's session
        set_session_value(session_id, "df_vendas", resultado["df_vendas"])
        set_session_value(session_id, "df_clientes", df_clientes)

        # Process IAF if available
        try:
            df_iaf = cruzar_vendas_com_iaf(resultado["df_vendas"], conn)
            set_session_value(session_id, "df_iaf", df_iaf)
            print(f"[Session {session_id[:8]}] IAF processado: {len(df_iaf)} registros")
        except Exception as e:
            print(f"[Session {session_id[:8]}] Erro ao processar IAF: {e}")
            set_session_value(session_id, "df_iaf", pl.DataFrame())

        response_data = {
            "success": True,
            "message": "Arquivo processado com sucesso",
            "estatisticas": resultado["estatisticas"],
            "avisos": resultado["avisos"]
        }

        return create_response_with_session(response_data, session_id)

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")


# =============================================================================
# FILTERS
# =============================================================================

@api_router.get("/filtros", response_model=FiltrosDisponiveis)
async def get_filtros(
    session: tuple = Depends(get_user_session),
):
    """Get available filter options from current data."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return FiltrosDisponiveis(ciclos=[], setores=[], marcas=[], gerencias=[])

    return FiltrosDisponiveis(
        ciclos=obter_ciclos_unicos(df_vendas),
        setores=obter_setores_unicos(df_vendas),
        marcas=obter_marcas_unicas(df_vendas),
        gerencias=obter_gerencias_unicas(df_vendas),
    )


# =============================================================================
# DASHBOARD METRICS
# =============================================================================

@api_router.get("/metricas/gerais")
async def get_metricas_gerais(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get general dashboard metrics."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")
    df_clientes = session_data.get("df_clientes")

    if df_vendas is None or df_clientes is None:
        return {"error": "Nenhum dado carregado. Faca upload de uma planilha."}

    # Apply filters
    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_vendas_filtrado = aplicar_filtros(df_vendas, ciclos=ciclos_list, setores=setores_list, gerencias=gerencias_list)
    df_clientes_filtrado = aplicar_filtros(df_clientes, ciclos=ciclos_list, setores=setores_list, gerencias=gerencias_list)

    metricas = calcular_metricas_gerais(df_clientes_filtrado, df_vendas_filtrado)
    return metricas


@api_router.get("/metricas/marcas")
async def get_vendas_por_marca(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get sales breakdown by brand."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return []

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_filtrado = aplicar_filtros(df_vendas, ciclos=ciclos_list, setores=setores_list, gerencias=gerencias_list)
    return calcular_vendas_por_marca(df_filtrado)


@api_router.get("/metricas/top-setores")
async def get_top_setores(
    limite: int = Query(5, ge=1, le=20),
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get top sectors by value."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")

    if df_clientes is None:
        return []

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None
    df_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )

    return calcular_top_setores(df_filtrado, limite=limite)


@api_router.get("/metricas/evolucao")
async def get_evolucao_ciclos(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get metrics evolution by cycle."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")

    if df_clientes is None:
        return []

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None
    df_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )

    return calcular_evolucao_ciclos(df_filtrado)


@api_router.get("/metricas/top10-setores")
async def get_top10_setores(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get top 10 sectors by value with complete data."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")

    if df_clientes is None:
        return []

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None
    df_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )

    return calcular_top_setores_completo(df_filtrado, limite=10)


@api_router.get("/metricas/resumo-ciclos")
async def get_resumo_ciclos(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get summary metrics by cycle."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")
    df_vendas = session_data.get("df_vendas")

    if df_clientes is None or df_vendas is None:
        return []

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_clientes_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )
    df_vendas_filtrado = aplicar_filtros(
        df_vendas,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )

    return calcular_resumo_ciclos(df_clientes_filtrado, df_vendas_filtrado)


@api_router.get("/metricas/dados-setor-ciclo")
async def get_dados_setor_ciclo(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get detailed data by sector and cycle including gerencia."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")
    df_vendas = session_data.get("df_vendas")

    if df_clientes is None or df_vendas is None:
        return []

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_clientes_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )
    df_vendas_filtrado = aplicar_filtros(
        df_vendas,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )

    return calcular_dados_setor_ciclo(df_clientes_filtrado, df_vendas_filtrado)


@api_router.get("/dashboard/export")
async def export_dashboard(
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    tabela: str = Query("todos", pattern="^(top10|resumo|setor_ciclo|todos)$"),
    session: tuple = Depends(get_user_session),
):
    """Export dashboard tables to CSV or Excel."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")
    df_vendas = session_data.get("df_vendas")

    if df_clientes is None or df_vendas is None:
        raise HTTPException(status_code=400, detail="Nenhum dado carregado")

    # Get data for each table
    top10_data = calcular_top_setores_completo(df_clientes, limite=10)
    resumo_data = calcular_resumo_ciclos(df_clientes, df_vendas)
    setor_ciclo_data = calcular_dados_setor_ciclo(df_clientes, df_vendas)

    # Format percentages for export
    resumo_data_formatted = [
        {**item, "percent_multimarcas": f"{item['percent_multimarcas']:.2f}%"}
        for item in resumo_data
    ] if resumo_data else []

    setor_ciclo_data_formatted = [
        {**item, "percent_multimarcas": f"{item['percent_multimarcas']:.2f}%"}
        for item in setor_ciclo_data
    ] if setor_ciclo_data else []

    if tabela == "top10":
        df = pl.DataFrame(top10_data)
        sheet_name = "Top10Setores"
        filename = "top10_setores"
    elif tabela == "resumo":
        df = pl.DataFrame(resumo_data_formatted)
        sheet_name = "ResumoCiclos"
        filename = "resumo_ciclos"
    elif tabela == "setor_ciclo":
        df = pl.DataFrame(setor_ciclo_data_formatted)
        sheet_name = "DadosSetorCiclo"
        filename = "dados_setor_ciclo"
    else:
        # Export all tables
        if formato == "xlsx":
            dataframes = {
                "Top10Setores": pl.DataFrame(top10_data) if top10_data else pl.DataFrame(),
                "ResumoCiclos": pl.DataFrame(resumo_data_formatted) if resumo_data_formatted else pl.DataFrame(),
                "DadosSetorCiclo": pl.DataFrame(setor_ciclo_data_formatted) if setor_ciclo_data_formatted else pl.DataFrame(),
            }
            content = exportar_multiplas_abas(dataframes)
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=dashboard_completo.xlsx"}
            )
        else:
            # For CSV, concatenate all data or export largest table
            df = pl.DataFrame(setor_ciclo_data_formatted) if setor_ciclo_data_formatted else pl.DataFrame()
            sheet_name = "Dashboard"
            filename = "dashboard"

    if df.is_empty():
        raise HTTPException(status_code=404, detail="Nenhum dado encontrado para exportar")

    if formato == "xlsx":
        content = exportar_excel(df, sheet_name)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{filename}.xlsx"
    else:
        content = exportar_csv(df)
        media_type = "text/csv"
        filename = f"{filename}.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# MULTIMARCAS
# =============================================================================

@api_router.get("/multimarcas")
async def get_multimarcas(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    limite: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    session: tuple = Depends(get_user_session),
):
    """Get multi-brand customers list."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")

    if df_clientes is None:
        return {"data": [], "total": 0}

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
        apenas_multimarcas=True
    )

    total = len(df_filtrado)
    df_paginado = df_filtrado.slice(offset, limite)

    data = [
        {
            "ciclo": row["CicloFaturamento"],
            "setor": row["Setor"],
            "codigo": row["CodigoRevendedora"],
            "nome": row["NomeRevendedora"],
            "qtde_marcas": row["QtdeMarcasDistintas"],
            "marcas": row["MarcasCompradas"],
            "itens": int(row["ItensTotal"] or 0),
            "valor": float(row["ValorTotal"] or 0),
        }
        for row in df_paginado.iter_rows(named=True)
    ]

    return {"data": data, "total": total}


@api_router.get("/multimarcas/combinacoes")
async def get_combinacoes_marcas(
    limite: int = Query(20, ge=1, le=50),
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get most frequent brand combinations."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")

    if df_clientes is None:
        return []

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None

    df_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        apenas_multimarcas=True
    )

    return calcular_combinacoes_marcas(df_filtrado, limite=limite)


@api_router.get("/multimarcas/export")
async def export_multimarcas(
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Export multi-brand customers to CSV or Excel."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")

    if df_clientes is None:
        raise HTTPException(status_code=400, detail="Nenhum dado carregado")

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None

    df_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        apenas_multimarcas=True
    )

    if formato == "xlsx":
        content = exportar_excel(df_filtrado, "Multimarcas")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "multimarcas.xlsx"
    else:
        content = exportar_csv(df_filtrado)
        media_type = "text/csv"
        filename = "multimarcas.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# CLIENTE
# =============================================================================

@api_router.get("/clientes")
async def listar_clientes(
    busca: Optional[str] = Query(None),
    limite: int = Query(50, ge=1, le=200),
    session: tuple = Depends(get_user_session),
):
    """List customers for selection."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")

    if df_clientes is None:
        return []

    df_unicos = df_clientes.select([
        "ClienteID",
        "NomeRevendedora",
        "CodigoRevendedora",
        "Setor"
    ]).unique()

    if busca:
        df_unicos = df_unicos.filter(
            pl.col("NomeRevendedora").str.to_lowercase().str.contains(busca.lower()) |
            pl.col("CodigoRevendedora").cast(pl.Utf8).str.contains(busca)
        )

    df_unicos = df_unicos.head(limite)

    return [
        {
            "cliente_id": row["ClienteID"],
            "nome": row["NomeRevendedora"],
            "codigo": row["CodigoRevendedora"],
            "setor": row["Setor"],
        }
        for row in df_unicos.iter_rows(named=True)
    ]


@api_router.get("/clientes/{cliente_id}")
async def get_cliente_detalhe(
    cliente_id: str,
    session: tuple = Depends(get_user_session),
):
    """Get detailed information for a specific customer."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        raise HTTPException(status_code=400, detail="Nenhum dado carregado")

    return obter_detalhes_cliente(df_vendas, cliente_id)


# =============================================================================
# AUDITORIA
# =============================================================================

@api_router.get("/auditoria/estatisticas")
async def get_auditoria_estatisticas(
    session: tuple = Depends(get_user_session),
):
    """Get audit statistics."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return {"error": "Nenhum dado carregado"}

    return obter_estatisticas_auditoria(df_vendas)


@api_router.get("/auditoria")
async def get_auditoria(
    motivo: Optional[str] = Query(None),
    limite: int = Query(100, ge=1, le=1000),
    session: tuple = Depends(get_user_session),
):
    """Get audit records list."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return []

    return listar_auditoria(df_vendas, motivo=motivo, limite=limite)


@api_router.get("/auditoria/export")
async def export_auditoria(
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    motivo: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Export audit records to CSV or Excel."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        raise HTTPException(status_code=400, detail="Nenhum dado carregado")

    auditoria_data = listar_auditoria(df_vendas, motivo=motivo, limite=5000)

    if not auditoria_data:
        raise HTTPException(status_code=404, detail="Nenhum registro de auditoria encontrado")

    df = pl.DataFrame(auditoria_data)

    if formato == "xlsx":
        content = exportar_excel(df, "Auditoria")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "auditoria.xlsx"
    else:
        content = exportar_csv(df)
        media_type = "text/csv"
        filename = "auditoria.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/produtos-novos")
async def get_produtos_novos(
    limite: int = Query(100, ge=1, le=500),
    session: tuple = Depends(get_user_session),
):
    """Get unregistered products list."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return []

    return listar_produtos_novos(df_vendas, limite=limite)


@api_router.get("/produtos-novos/export")
async def export_produtos_novos(
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    session: tuple = Depends(get_user_session),
):
    """Export unregistered products to CSV or Excel."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        raise HTTPException(status_code=400, detail="Nenhum dado carregado")

    from app.services.auditoria import gerar_produtos_nao_cadastrados
    df_novos = gerar_produtos_nao_cadastrados(df_vendas)

    if df_novos.is_empty():
        raise HTTPException(status_code=404, detail="Nenhum produto novo encontrado")

    if formato == "xlsx":
        content = exportar_excel(df_novos, "ProdutosNovos")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "produtos_novos.xlsx"
    else:
        content = exportar_csv(df_novos)
        media_type = "text/csv"
        filename = "produtos_novos.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# IAF
# =============================================================================

@api_router.get("/iaf/metricas")
async def get_iaf_metricas(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get IAF penetration metrics."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")
    df_iaf = session_data.get("df_iaf")

    if df_clientes is None:
        return {"error": "Nenhum dado carregado"}

    if df_iaf is None:
        df_iaf = pl.DataFrame()

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_clientes_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )
    df_iaf_filtrado = aplicar_filtros(
        df_iaf,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )

    return {
        "cabelos": calcular_percentual_iaf(df_clientes_filtrado, df_iaf_filtrado, "Cabelos"),
        "make": calcular_percentual_iaf(df_clientes_filtrado, df_iaf_filtrado, "Make"),
        "total": calcular_percentual_iaf(df_clientes_filtrado, df_iaf_filtrado, None),
    }


@api_router.get("/iaf/por-setor")
async def get_iaf_por_setor(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get IAF metrics by sector."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")
    df_iaf = session_data.get("df_iaf")

    if df_clientes is None:
        return []

    if df_iaf is None:
        df_iaf = pl.DataFrame()

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_clientes_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )
    df_iaf_filtrado = aplicar_filtros(
        df_iaf,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )

    return calcular_iaf_por_setor(df_clientes_filtrado, df_iaf_filtrado)


# =============================================================================
# METAS POR SETOR
# =============================================================================

@api_router.get("/metas/por-setor")
async def get_metas_por_setor(
    ciclos: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Get per-sector metrics for goal tracking page."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")
    df_vendas = session_data.get("df_vendas")
    df_iaf = session_data.get("df_iaf")

    if df_clientes is None or df_vendas is None:
        return []

    if df_iaf is None:
        df_iaf = pl.DataFrame()

    ciclos_list = ciclos.split(",") if ciclos else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_clientes_f = aplicar_filtros(df_clientes, ciclos=ciclos_list, gerencias=gerencias_list)
    df_vendas_f = aplicar_filtros(df_vendas, ciclos=ciclos_list, gerencias=gerencias_list)
    df_iaf_f = aplicar_filtros(df_iaf, ciclos=ciclos_list, gerencias=gerencias_list)

    metricas = calcular_metricas_por_setor(df_clientes_f)
    iaf_por_setor = calcular_iaf_por_setor(df_clientes_f, df_iaf_f)
    iaf_dict = {item["setor"]: item for item in iaf_por_setor}

    metas_planilha = ler_planilha_metas()

    for m in metricas:
        iaf = iaf_dict.get(m["setor"], {})
        m["clientes_cabelos"] = iaf.get("clientes_cabelos", 0)
        m["percent_cabelos"] = iaf.get("percent_cabelos", 0.0)
        m["clientes_make"] = iaf.get("clientes_make", 0)
        m["percent_make"] = iaf.get("percent_make", 0.0)

        meta_p = encontrar_meta_setor(m["setor"], metas_planilha)
        if meta_p:
            m["supervisora"] = meta_p["supervisora"]
            m["meta_planilha"] = {
                "receita":             meta_p["receita"],
                "clientes_ativos":     meta_p["ativo"],
                "rpa":                 meta_p["rpa"],
                "percent_multimarcas": meta_p["multimarca_pct"],
                "clientes_multimarcas":meta_p["multimarca_qtd"],
                "percent_cabelos":     meta_p["cabelo_pct"],
                "clientes_cabelos":    meta_p["cabelo_qtd"],
                "percent_make":        meta_p["make_pct"],
                "clientes_make":       meta_p["make_qtd"],
            }
        else:
            m["supervisora"] = ""
            m["meta_planilha"] = None

    return metricas


@api_router.get("/metas/planilha")
async def get_metas_planilha():
    """Return the raw parsed content of metas.xlsx."""
    return ler_planilha_metas()


@api_router.get("/iaf/vendas")
async def get_iaf_vendas(
    tipo: Optional[str] = Query(None),
    setor: Optional[str] = Query(None),
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    limite: int = Query(200, ge=1, le=500),
    session: tuple = Depends(get_user_session),
):
    """Get IAF sales list."""
    session_id, session_data = session
    df_iaf = session_data.get("df_iaf")

    if df_iaf is None:
        return []

    if isinstance(df_iaf, pl.DataFrame) and df_iaf.is_empty():
        return []

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    if setor:
        setores_list = [setor]

    df_iaf_filtrado = aplicar_filtros(
        df_iaf,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )

    return listar_vendas_iaf(df_iaf_filtrado, tipo_iaf=tipo, setor=None, limite=limite)


@api_router.get("/iaf/export")
async def export_iaf(
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    tabela: str = Query("setor", pattern="^(setor|vendas|todos)$"),
    tipo: Optional[str] = Query(None),
    setor: Optional[str] = Query(None),
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Export IAF data to CSV or Excel."""
    session_id, session_data = session
    df_clientes = session_data.get("df_clientes")
    df_iaf = session_data.get("df_iaf")

    if df_clientes is None:
        raise HTTPException(status_code=400, detail="Nenhum dado carregado")

    if df_iaf is None:
        df_iaf = pl.DataFrame()

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    if setor:
        setores_list = [setor]

    df_clientes_filtrado = aplicar_filtros(
        df_clientes,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )
    df_iaf_filtrado = aplicar_filtros(
        df_iaf,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list,
    )

    # Get data
    setor_data = calcular_iaf_por_setor(df_clientes_filtrado, df_iaf_filtrado)
    vendas_data = listar_vendas_iaf(df_iaf_filtrado, tipo_iaf=tipo, setor=None, limite=1000) if not df_iaf_filtrado.is_empty() else []

    # Format percentages for export
    setor_data_formatted = [
        {
            **item,
            "percent_iaf": f"{item['percent_iaf']}%",
            "percent_cabelos": f"{item['percent_cabelos']}%",
            "percent_make": f"{item['percent_make']}%",
        }
        for item in setor_data
    ] if setor_data else []

    if tabela == "setor":
        df = pl.DataFrame(setor_data_formatted) if setor_data_formatted else pl.DataFrame()
        sheet_name = "IAFPorSetor"
        filename = "iaf_por_setor"
    elif tabela == "vendas":
        df = pl.DataFrame(vendas_data) if vendas_data else pl.DataFrame()
        sheet_name = "VendasIAF"
        filename = "vendas_iaf"
    else:
        # Export all tables
        if formato == "xlsx":
            dataframes = {
                "IAFPorSetor": pl.DataFrame(setor_data_formatted) if setor_data_formatted else pl.DataFrame(),
                "VendasIAF": pl.DataFrame(vendas_data) if vendas_data else pl.DataFrame(),
            }
            content = exportar_multiplas_abas(dataframes)
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=iaf_completo.xlsx"}
            )
        else:
            df = pl.DataFrame(setor_data_formatted) if setor_data_formatted else pl.DataFrame()
            sheet_name = "IAF"
            filename = "iaf"

    if df.is_empty():
        raise HTTPException(status_code=404, detail="Nenhum dado encontrado para exportar")

    if formato == "xlsx":
        content = exportar_excel(df, sheet_name)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{filename}.xlsx"
    else:
        content = exportar_csv(df)
        media_type = "text/csv"
        filename = f"{filename}.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# CATEGORIAS
# =============================================================================

@api_router.get("/categorias/lista")
async def get_categorias_lista():
    """Get list of available categories."""
    return obter_categorias_disponiveis()


@api_router.get("/categorias/metricas")
async def get_categorias_metricas(
    session: tuple = Depends(get_user_session),
):
    """Get metrics by category."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return []

    # Classify and calculate metrics
    df_classificado = classificar_vendas(df_vendas)
    return calcular_metricas_categoria(df_classificado)


@api_router.get("/categorias/por-ciclo")
async def get_categorias_por_ciclo(
    session: tuple = Depends(get_user_session),
):
    """Get category metrics by cycle."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return []

    df_classificado = classificar_vendas(df_vendas)
    return calcular_categoria_por_ciclo(df_classificado)


@api_router.get("/categorias/por-setor")
async def get_categorias_por_setor(
    session: tuple = Depends(get_user_session),
):
    """Get category metrics by sector."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return []

    df_classificado = classificar_vendas(df_vendas)
    return calcular_categoria_por_setor(df_classificado)


@api_router.get("/categorias/{categoria}/produtos")
async def get_produtos_categoria(
    categoria: str,
    limite: int = Query(50, ge=1, le=200),
    session: tuple = Depends(get_user_session),
):
    """Get products in a specific category."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return []

    df_classificado = classificar_vendas(df_vendas)
    return listar_produtos_categoria(df_classificado, categoria, limite=limite)


@api_router.get("/categorias/export")
async def export_categorias(
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    tabela: str = Query("metricas", pattern="^(metricas|ciclo|setor|todos)$"),
    session: tuple = Depends(get_user_session),
):
    """Export category data to CSV or Excel."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        raise HTTPException(status_code=400, detail="Nenhum dado carregado")

    df_classificado = classificar_vendas(df_vendas)

    metricas_data = calcular_metricas_categoria(df_classificado)
    ciclo_data = calcular_categoria_por_ciclo(df_classificado)
    setor_data = calcular_categoria_por_setor(df_classificado)

    # Format percentages for export
    metricas_data_formatted = [
        {
            **item,
            "percent_valor": f"{item['percent_valor']}%",
            "percent_itens": f"{item['percent_itens']}%",
        }
        for item in metricas_data
    ] if metricas_data else []

    if tabela == "metricas":
        df = pl.DataFrame(metricas_data_formatted) if metricas_data_formatted else pl.DataFrame()
        sheet_name = "MetricasCategorias"
        filename = "categorias_metricas"
    elif tabela == "ciclo":
        df = pl.DataFrame(ciclo_data) if ciclo_data else pl.DataFrame()
        sheet_name = "CategoriasPorCiclo"
        filename = "categorias_por_ciclo"
    elif tabela == "setor":
        df = pl.DataFrame(setor_data) if setor_data else pl.DataFrame()
        sheet_name = "CategoriasPorSetor"
        filename = "categorias_por_setor"
    else:
        # Export all tables
        if formato == "xlsx":
            dataframes = {
                "Metricas": pl.DataFrame(metricas_data_formatted) if metricas_data_formatted else pl.DataFrame(),
                "PorCiclo": pl.DataFrame(ciclo_data) if ciclo_data else pl.DataFrame(),
                "PorSetor": pl.DataFrame(setor_data) if setor_data else pl.DataFrame(),
            }
            content = exportar_multiplas_abas(dataframes)
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=categorias_completo.xlsx"}
            )
        else:
            df = pl.DataFrame(metricas_data_formatted) if metricas_data_formatted else pl.DataFrame()
            sheet_name = "Categorias"
            filename = "categorias"

    if df.is_empty():
        raise HTTPException(status_code=404, detail="Nenhum dado encontrado para exportar")

    if formato == "xlsx":
        content = exportar_excel(df, sheet_name)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{filename}.xlsx"
    else:
        content = exportar_csv(df)
        media_type = "text/csv"
        filename = f"{filename}.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# RANKING DE REVENDEDORAS
# =============================================================================

@api_router.get("/ranking/revendedoras")
async def get_ranking_revendedoras(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    limite: int = Query(20, ge=1, le=100),
    session: tuple = Depends(get_user_session),
):
    """Get top resellers ranking by value."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return []

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_filtrado = aplicar_filtros(
        df_vendas,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list
    )

    return calcular_ranking_revendedoras(df_filtrado, limite=limite)


@api_router.get("/ranking/revendedora/{codigo}/evolucao")
async def get_evolucao_revendedora(
    codigo: str,
    session: tuple = Depends(get_user_session),
):
    """Get a reseller's evolution over cycles."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        return []

    return calcular_evolucao_revendedora(df_vendas, codigo)


@api_router.get("/ranking/export")
async def export_ranking(
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    limite: int = Query(100, ge=1, le=500),
    session: tuple = Depends(get_user_session),
):
    """Export ranking revendedoras to CSV or Excel."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")

    if df_vendas is None:
        raise HTTPException(status_code=400, detail="Nenhum dado carregado")

    ciclos_list = ciclos.split(",") if ciclos else None
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_filtrado = aplicar_filtros(
        df_vendas,
        ciclos=ciclos_list,
        setores=setores_list,
        gerencias=gerencias_list
    )

    ranking_data = calcular_ranking_revendedoras(df_filtrado, limite=limite)

    if not ranking_data:
        raise HTTPException(status_code=404, detail="Nenhum dado encontrado")

    df = pl.DataFrame(ranking_data)

    if formato == "xlsx":
        content = exportar_excel(df, "RankingRevendedoras")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "ranking_revendedoras.xlsx"
    else:
        content = exportar_csv(df)
        media_type = "text/csv"
        filename = "ranking_revendedoras.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# COMPARATIVO DE CICLOS
# =============================================================================

@api_router.get("/comparativo/ciclos")
async def get_comparativo_ciclos(
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """
    Get comparison metrics between selected cycles.

    - 1 cycle: Returns normal metrics
    - 2+ cycles: Returns metrics with variations between cycles
    """
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")
    df_clientes = session_data.get("df_clientes")

    if df_vendas is None or df_clientes is None:
        return {"ciclos": [], "metricas": [], "total_ciclos": 0}

    # Get cycles to compare
    if ciclos:
        ciclos_list = ciclos.split(",")
    else:
        # If no cycles specified, get all available
        from app.services.venda import obter_ciclos_unicos
        ciclos_list = obter_ciclos_unicos(df_vendas)

    # Apply other filters
    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_vendas_filtrado = aplicar_filtros(
        df_vendas,
        setores=setores_list,
        gerencias=gerencias_list
    )
    df_clientes_filtrado = aplicar_filtros(
        df_clientes,
        setores=setores_list,
        gerencias=gerencias_list
    )

    return calcular_comparativo_ciclos(df_clientes_filtrado, df_vendas_filtrado, ciclos_list)


@api_router.get("/comparativo/export")
async def export_comparativo(
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    ciclos: Optional[str] = Query(None),
    setores: Optional[str] = Query(None),
    gerencias: Optional[str] = Query(None),
    session: tuple = Depends(get_user_session),
):
    """Export cycle comparison to CSV or Excel."""
    session_id, session_data = session
    df_vendas = session_data.get("df_vendas")
    df_clientes = session_data.get("df_clientes")

    if df_vendas is None or df_clientes is None:
        raise HTTPException(status_code=400, detail="Nenhum dado carregado")

    # Get cycles to compare
    if ciclos:
        ciclos_list = ciclos.split(",")
    else:
        from app.services.venda import obter_ciclos_unicos
        ciclos_list = obter_ciclos_unicos(df_vendas)

    setores_list = setores.split(",") if setores else None
    gerencias_list = gerencias.split(",") if gerencias else None

    df_vendas_filtrado = aplicar_filtros(
        df_vendas,
        setores=setores_list,
        gerencias=gerencias_list
    )
    df_clientes_filtrado = aplicar_filtros(
        df_clientes,
        setores=setores_list,
        gerencias=gerencias_list
    )

    comparativo = calcular_comparativo_ciclos(df_clientes_filtrado, df_vendas_filtrado, ciclos_list)

    if not comparativo.get("metricas"):
        raise HTTPException(status_code=404, detail="Nenhum dado encontrado")

    df = pl.DataFrame(comparativo["metricas"])

    if formato == "xlsx":
        content = exportar_excel(df, "ComparativoCiclos")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "comparativo_ciclos.xlsx"
    else:
        content = exportar_csv(df)
        media_type = "text/csv"
        filename = "comparativo_ciclos.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# PRODUCT MANAGEMENT ENDPOINTS
# =============================================================================

@api_router.get("/marcas-disponiveis")
async def get_marcas_disponiveis():
    """Get list of available brands for product registration."""
    return MARCAS_GRUPO


@api_router.post("/produtos/cadastrar")
async def cadastrar_produto(
    sku: str = Form(...),
    nome: str = Form(...),
    marca: str = Form(...),
    conn: sqlite3.Connection = Depends(get_db),
    session: tuple = Depends(get_user_session),
):
    """
    Register a new product in the database.

    Args:
        sku: Product SKU code
        nome: Product name
        marca: Brand name (must be one of MARCAS_GRUPO)

    Returns:
        Success message with product details
    """
    # Validate brand
    if marca not in MARCAS_GRUPO:
        raise HTTPException(
            status_code=400,
            detail=f"Marca invalida. Opcoes: {', '.join(MARCAS_GRUPO)}"
        )

    # Normalize SKU
    sku_normalizado = normalizar_sku(sku)
    if not sku_normalizado:
        raise HTTPException(status_code=400, detail="SKU invalido")

    # Check if already exists
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id FROM produtos WHERE sku_normalizado = ?",
        (sku_normalizado,)
    )
    if cursor.fetchone():
        raise HTTPException(
            status_code=400,
            detail=f"Produto com SKU {sku} ja existe no banco de dados"
        )

    # Insert product
    try:
        session_id, session_data = session

        cursor.execute(
            """
            INSERT INTO produtos (sku, sku_normalizado, nome, marca)
            VALUES (?, ?, ?, ?)
            """,
            (sku, sku_normalizado, nome.strip(), marca)
        )
        _upsert_produtos_na_planilha([{
            "sku": sku,
            "nome": nome.strip(),
            "marca": marca,
        }])
        threading.Thread(target=_backup_para_github, daemon=True).start()
        conn.commit()
        linhas_sessao_atualizadas = _atualizar_sessao_com_produtos_cadastrados(
            session_id,
            session_data,
            [{"sku": sku, "nome": nome.strip(), "marca": marca}],
        )

        return {
            "success": True,
            "message": f"Produto cadastrado com sucesso",
            "produto": {
                "sku": sku,
                "sku_normalizado": sku_normalizado,
                "nome": nome.strip(),
                "marca": marca
            },
            "sessao_atualizada": linhas_sessao_atualizadas > 0,
            "linhas_sessao_atualizadas": linhas_sessao_atualizadas,
        }
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao cadastrar: {str(e)}")


@api_router.post("/produtos/cadastrar-lote")
async def cadastrar_produtos_lote(
    produtos: List[Dict[str, str]],
    conn: sqlite3.Connection = Depends(get_db),
    session: tuple = Depends(get_user_session),
):
    """
    Register multiple products at once.

    Args:
        produtos: List of dicts with sku, nome, marca

    Returns:
        Summary of registered products
    """
    cursor = conn.cursor()
    cadastrados = 0
    erros = []

    produtos_para_planilha: List[Dict[str, str]] = []

    for p in produtos:
        sku = p.get("sku", "")
        nome = p.get("nome", "")
        marca = p.get("marca", "")

        if marca not in MARCAS_GRUPO:
            erros.append(f"SKU {sku}: marca invalida '{marca}'")
            continue

        sku_normalizado = normalizar_sku(sku)
        if not sku_normalizado:
            erros.append(f"SKU {sku}: SKU invalido")
            continue

        # Check if exists
        cursor.execute(
            "SELECT id FROM produtos WHERE sku_normalizado = ?",
            (sku_normalizado,)
        )
        if cursor.fetchone():
            erros.append(f"SKU {sku}: ja existe")
            continue

        try:
            cursor.execute(
                """
                INSERT INTO produtos (sku, sku_normalizado, nome, marca)
                VALUES (?, ?, ?, ?)
                """,
                (sku, sku_normalizado, nome.strip(), marca)
            )
            cadastrados += 1
            produtos_para_planilha.append({
                "sku": sku,
                "nome": nome.strip(),
                "marca": marca,
            })
        except Exception as e:
            erros.append(f"SKU {sku}: {str(e)}")

    try:
        if produtos_para_planilha:
            _upsert_produtos_na_planilha(produtos_para_planilha)
            threading.Thread(target=_backup_para_github, daemon=True).start()
        conn.commit()
        session_id, session_data = session
        linhas_sessao_atualizadas = _atualizar_sessao_com_produtos_cadastrados(
            session_id,
            session_data,
            produtos_para_planilha,
        )
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao salvar lote na planilha BD: {str(e)}")

    return {
        "success": True,
        "cadastrados": cadastrados,
        "erros": erros,
        "total_erros": len(erros),
        "sessao_atualizada": linhas_sessao_atualizadas > 0,
        "linhas_sessao_atualizadas": linhas_sessao_atualizadas,
    }


# =============================================================================
# GEOGRAPHIC ANALYSIS – UPLOAD
# =============================================================================

@api_router.post("/upload-clientes")
async def upload_clientes(
    request: Request,
    file: UploadFile = File(...),
):
    """
    Upload and process the clients spreadsheet for geographic analysis.

    Persists the processed DataFrame to disk (Parquet) so it survives
    server restarts.  Also stores it in app.state for instant access.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo não fornecido")

    if not file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Formato inválido. Use CSV ou Excel (.xlsx, .xls)",
        )

    try:
        content = await file.read()
        resultado = processar_planilha_clientes(content, file.filename)

        # Persist to disk
        resultado["df"].write_parquet(GEO_PARQUET_PATH)
        with open(GEO_STATS_PATH, "w", encoding="utf-8") as fh:
            _json.dump(resultado["estatisticas"], fh, ensure_ascii=False)

        # Store in app-level state (shared across all sessions)
        request.app.state.df_geo       = resultado["df"]
        request.app.state.df_geo_stats = resultado["estatisticas"]

        return {
            "success": True,
            "message": "Planilha de clientes processada com sucesso",
            "estatisticas": resultado["estatisticas"],
            "avisos": resultado["avisos"],
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")


# =============================================================================
# GEOGRAPHIC ANALYSIS – ENDPOINTS
# =============================================================================

@api_router.get("/geo/status")
async def geo_status(request: Request):
    """Check whether the clients geographic spreadsheet has been loaded."""
    df_geo = getattr(request.app.state, "df_geo", None)
    stats  = getattr(request.app.state, "df_geo_stats", {}) or {}
    return {
        "has_data": df_geo is not None and len(df_geo) > 0,
        "estatisticas": stats,
    }


@api_router.get("/geo/bairros")
async def get_geo_bairros(
    request: Request,
    unidade: Optional[str] = Query(None),
    cidade: Optional[str] = Query(None),
    situacao: Optional[str] = Query(None),
):
    """Return client metrics grouped by neighborhood."""
    df_geo = getattr(request.app.state, "df_geo", None)
    if df_geo is None or len(df_geo) == 0:
        return {"bairros": []}
    return {
        "bairros": calcular_metricas_bairro(
            df_geo,
            unidade=unidade or None,
            cidade=cidade or None,
            situacao=situacao or None,
        )
    }


@api_router.get("/geo/cidades")
async def get_geo_cidades(
    request: Request,
    unidade: Optional[str] = Query(None),
):
    """Return client counts per city (used to color the heat map)."""
    df_geo = getattr(request.app.state, "df_geo", None)
    if df_geo is None or len(df_geo) == 0:
        return {"cidades": []}
    return {
        "cidades": calcular_metricas_cidade(df_geo, unidade=unidade or None)
    }


@api_router.get("/geo/clientes")
async def get_geo_clientes(
    request: Request,
    unidade: Optional[str] = Query(None),
    cidade: Optional[str] = Query(None),
    bairro: Optional[str] = Query(None),
    situacao: Optional[str] = Query(None),
    ordenar_por: str = Query("ciclos_desc"),
):
    """Return individual clients with geographic and inactivity data."""
    df_geo = getattr(request.app.state, "df_geo", None)
    if df_geo is None or len(df_geo) == 0:
        return {"clientes": []}
    return {
        "clientes": listar_clientes_geo(
            df_geo,
            unidade=unidade or None,
            cidade=cidade or None,
            bairro=bairro or None,
            situacao=situacao or None,
            ordenar_por=ordenar_por,
        )
    }


@api_router.get("/geo/filtros")
async def get_geo_filtros(
    request: Request,
    unidade: Optional[str] = Query(None),
):
    """Return available cities and neighborhoods for filter dropdowns."""
    df_geo = getattr(request.app.state, "df_geo", None)
    if df_geo is None or len(df_geo) == 0:
        return {"cidades": [], "bairros": []}

    from app.services.geo import _aplicar_filtros
    df_filtered = _aplicar_filtros(df_geo, unidade=unidade or None)
    return {
        "cidades": obter_cidades_geo(df_filtered),
        "bairros": obter_bairros_geo(df_filtered),
    }


@api_router.get("/geo/export/bairros")
async def export_geo_bairros(
    request: Request,
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    unidade: Optional[str] = Query(None),
    cidade: Optional[str] = Query(None),
    situacao: Optional[str] = Query(None),
):
    """Export neighborhood analysis as CSV or Excel."""
    df_geo = getattr(request.app.state, "df_geo", None)
    if df_geo is None or len(df_geo) == 0:
        raise HTTPException(status_code=400, detail="Sem dados para exportar")

    rows = calcular_metricas_bairro(
        df_geo,
        unidade=unidade or None,
        cidade=cidade or None,
        situacao=situacao or None,
    )
    df_export = pl.DataFrame(rows)

    if formato == "csv":
        return Response(
            content=exportar_csv(df_export),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=analise_bairros.csv"},
        )
    else:
        return Response(
            content=exportar_excel(df_export, "Bairros"),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=analise_bairros.xlsx"},
        )


@api_router.get("/geo/bairro/detalhe")
async def get_geo_bairro_detalhe(
    request: Request,
    bairro: str = Query(...),
    cidade: Optional[str] = Query(None),
    unidade: Optional[str] = Query(None),
    situacao: Optional[str] = Query(None),
):
    """Return streets and clients for a specific neighborhood (accordion detail)."""
    df_geo = getattr(request.app.state, "df_geo", None)
    if df_geo is None or len(df_geo) == 0:
        return {"ruas": [], "clientes": []}
    return calcular_detalhe_bairro(
        df_geo,
        bairro=bairro,
        cidade=cidade or None,
        unidade=unidade or None,
        situacao=situacao or None,
    )


# IBGE Censo 2022 — população residente por município de Alagoas
# Fonte: SIDRA tabela 9514, variável 93, período 2022
# Embutido diretamente para evitar dependência de rede em runtime.
_POPULACAO_AL: dict = {
    "Água Branca": 19008,
    "Anadia": 13966,
    "Arapiraca": 234696,
    "Atalaia": 37512,
    "Barra de Santo Antônio": 16365,
    "Barra de São Miguel": 7944,
    "Batalha": 16448,
    "Belo Monte": 5936,
    "Belém": 4722,
    "Boca da Mata": 21187,
    "Branquinha": 9603,
    "Cacimbinhas": 10482,
    "Cajueiro": 16024,
    "Campestre": 6665,
    "Campo Alegre": 32106,
    "Campo Grande": 8143,
    "Canapi": 15559,
    "Capela": 15032,
    "Carneiros": 8999,
    "Chã Preta": 5910,
    "Coité do Nóia": 10810,
    "Colônia Leopoldina": 15816,
    "Coqueiro Seco": 5581,
    "Coruripe": 50414,
    "Craíbas": 25397,
    "Delmiro Gouveia": 51318,
    "Dois Riachos": 9805,
    "Estrela de Alagoas": 15429,
    "Feira Grande": 22712,
    "Feliz Deserto": 3963,
    "Flexeiras": 9618,
    "Girau do Ponciano": 36102,
    "Ibateguara": 13731,
    "Igaci": 23995,
    "Igreja Nova": 21372,
    "Inhapi": 15167,
    "Jacaré dos Homens": 5083,
    "Jacuípe": 5352,
    "Japaratinga": 9219,
    "Jaramataia": 4985,
    "Jequiá da Praia": 9470,
    "Joaquim Gomes": 17150,
    "Jundiá": 4092,
    "Junqueiro": 23907,
    "Lagoa da Canoa": 18457,
    "Limoeiro de Anadia": 24740,
    "Maceió": 957916,
    "Major Isidoro": 17700,
    "Mar Vermelho": 3155,
    "Maragogi": 32174,
    "Maravilha": 9534,
    "Marechal Deodoro": 60370,
    "Maribondo": 13679,
    "Mata Grande": 21844,
    "Matriz de Camaragibe": 23857,
    "Messias": 15405,
    "Minador do Negrão": 4845,
    "Monteirópolis": 7184,
    "Murici": 25187,
    "Novo Lino": 10020,
    "Olho d'Água Grande": 4330,
    "Olho d'Água das Flores": 20695,
    "Olho d'Água do Casado": 8349,
    "Olivença": 10812,
    "Ouro Branco": 11446,
    "Palestina": 4325,
    "Palmeira dos Índios": 71574,
    "Pariconha": 10573,
    "Paripueira": 13835,
    "Passo de Camaragibe": 13804,
    "Paulo Jacinto": 6576,
    "Penedo": 58650,
    "Piaçabuçu": 15908,
    "Pilar": 35370,
    "Pindoba": 2731,
    "Piranhas": 22609,
    "Porto Calvo": 24071,
    "Porto Real do Colégio": 20082,
    "Porto de Pedras": 9295,
    "Poço das Trincheiras": 12518,
    "Pão de Açúcar": 23823,
    "Quebrangulo": 11080,
    "Rio Largo": 93927,
    "Roteiro": 6474,
    "Santa Luzia do Norte": 6919,
    "Santana do Ipanema": 46220,
    "Santana do Mundaú": 11323,
    "Satuba": 24278,
    "Senador Rui Palmeira": 12303,
    "São Brás": 6555,
    "São José da Laje": 20813,
    "São José da Tapera": 30604,
    "São Luís do Quitunde": 30873,
    "São Miguel dos Campos": 51990,
    "São Miguel dos Milagres": 8482,
    "São Sebastião": 31786,
    "Tanque d'Arca": 5796,
    "Taquarana": 19032,
    "Teotônio Vilela": 38053,
    "Traipu": 23565,
    "União dos Palmares": 59280,
    "Viçosa": 24092,
}


@api_router.get("/geo/populacao")
async def geo_populacao():
    """Return IBGE Censo 2022 resident population for all Alagoas municipalities."""
    return {"municipios": _POPULACAO_AL}


@api_router.post("/geo/clear")
async def clear_geo(request: Request):
    """Remove the geographic clients data from disk and app state."""
    request.app.state.df_geo       = None
    request.app.state.df_geo_stats = {}
    for path in (GEO_PARQUET_PATH, GEO_STATS_PATH):
        try:
            os.remove(path)
        except FileNotFoundError:
            pass
    return {"success": True}


@api_router.get("/geo/export/clientes")
async def export_geo_clientes(
    request: Request,
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    unidade: Optional[str] = Query(None),
    cidade: Optional[str] = Query(None),
    bairro: Optional[str] = Query(None),
    situacao: Optional[str] = Query(None),
):
    """Export clients list as CSV or Excel."""
    df_geo = getattr(request.app.state, "df_geo", None)
    if df_geo is None or len(df_geo) == 0:
        raise HTTPException(status_code=400, detail="Sem dados para exportar")

    rows = listar_clientes_geo(
        df_geo,
        unidade=unidade or None,
        cidade=cidade or None,
        bairro=bairro or None,
        situacao=situacao or None,
        ordenar_por="ciclos_desc",
        limite=10000,
    )
    df_export = pl.DataFrame(rows)

    if formato == "csv":
        return Response(
            content=exportar_csv(df_export),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=clientes_geografico.csv"},
        )
    else:
        return Response(
            content=exportar_excel(df_export, "Clientes"),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=clientes_geografico.xlsx"},
        )

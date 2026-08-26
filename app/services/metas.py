"""
Service for reading and matching sector goals from metas.xlsx.

Sector names are normalized before comparison:
- uppercase
- trailing slashes/spaces removed
- spaces around '/' standardized to ' / '
- multiple spaces collapsed
"""
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from app.config import BASE_DIR

METAS_PATH = BASE_DIR / "metas.xlsx"

# Column names in metas.xlsx
_COL_SETOR        = "SETOR"
_COL_SUPERVISORA  = "SUPERVISORA"
_COL_RECEITA      = "RECEITA"
_COL_ATIVO        = "ATIVO"
_COL_RPA          = "RPA"
_COL_MULTI_PCT    = "MULTIMARCA (%)"
_COL_MULTI_QTD    = "MULTIMARCA (Qtd)"
_COL_CABELO_PCT   = "CABELO (%)"
_COL_CABELO_QTD   = "CABELO (Qtd)"
_COL_MAKE_PCT     = "MAKE (%)"
_COL_MAKE_QTD     = "MAKE (Qtd)"


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _parse_brl(value: str) -> float:
    """Parse Brazilian currency string 'R$ 50.000,00' → 50000.0."""
    if not value:
        return 0.0
    cleaned = str(value).replace("R$", "").replace(" ", "").strip()
    # Remove thousands separator (.) then swap decimal separator (, → .)
    cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _parse_pct(value: str) -> float:
    """Parse '0.73' → 73.0  or  '73' → 73.0  (always returns percentage points)."""
    try:
        v = float(str(value).strip().replace(",", "."))
        # Values ≤ 1.0 are fractions (0.73 = 73%)
        return round(v * 100, 1) if v <= 1.0 else round(v, 1)
    except (ValueError, TypeError):
        return 0.0


def _parse_int(value: str) -> int:
    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        return 0


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def ler_planilha_metas() -> List[Dict[str, Any]]:
    """
    Read metas.xlsx and return a list of parsed meta dicts.

    Uses openpyxl so that numeric cells are returned as floats (no need to
    parse currency strings for those columns) and accented names are read
    correctly as Unicode strings.

    Returns [] if the file does not exist or cannot be read.
    Each dict has keys:
        setor, supervisora,
        receita, ativo, rpa,
        multimarca_pct, multimarca_qtd,
        cabelo_pct, cabelo_qtd,
        make_pct, make_qtd
    """
    if not METAS_PATH.exists():
        return []

    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(METAS_PATH), data_only=True)
        ws = wb.active
    except Exception as exc:
        print(f"[WARN] Could not read metas.xlsx: {exc}")
        return []

    # Build column index from header row
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(headers)}

    def _get(row_vals: tuple, col_name: str):
        i = col_idx.get(col_name)
        return row_vals[i] if i is not None else None

    result: List[Dict[str, Any]] = []
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        setor = str(_get(row_vals, _COL_SETOR) or "").strip()
        if not setor:
            continue

        raw_receita = _get(row_vals, _COL_RECEITA)
        raw_rpa     = _get(row_vals, _COL_RPA)

        result.append(
            {
                "setor":          setor,
                "supervisora":    str(_get(row_vals, _COL_SUPERVISORA) or "").strip(),
                # RECEITA / RPA may be strings ("R$ 50.000,00") or numbers
                "receita":        _parse_brl(str(raw_receita)) if isinstance(raw_receita, str) else float(raw_receita or 0),
                "ativo":          _parse_int(str(_get(row_vals, _COL_ATIVO) or 0)),
                "rpa":            _parse_brl(str(raw_rpa)) if isinstance(raw_rpa, str) else float(raw_rpa or 0),
                # Percentages come as floats like 0.73 → convert to 73.0
                "multimarca_pct": _parse_pct(str(_get(row_vals, _COL_MULTI_PCT) or 0)),
                "multimarca_qtd": _parse_int(str(_get(row_vals, _COL_MULTI_QTD) or 0)),
                "cabelo_pct":     _parse_pct(str(_get(row_vals, _COL_CABELO_PCT) or 0)),
                "cabelo_qtd":     _parse_int(str(_get(row_vals, _COL_CABELO_QTD) or 0)),
                "make_pct":       _parse_pct(str(_get(row_vals, _COL_MAKE_PCT) or 0)),
                "make_qtd":       _parse_int(str(_get(row_vals, _COL_MAKE_QTD) or 0)),
            }
        )
    return result


def _normalizar(nome: str) -> str:
    """
    Normalize a sector name for comparison.

    Steps:
    - Uppercase
    - Strip leading/trailing whitespace and trailing slashes
    - Standardize spacing around '/' → ' / '
    - Collapse multiple spaces
    """
    nome = str(nome or "").upper().strip().rstrip("/ ").strip()
    nome = re.sub(r"\s*/\s*", " / ", nome)
    nome = re.sub(r"\s+", " ", nome)
    return nome


def encontrar_meta_setor(
    nome_setor_app: str,
    metas: List[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    """
    Find the planilha meta row whose key matches the app's sector name.

    Both sides are normalized before comparison (see _normalizar).
    Matching rules (in order):
    1. Igualdade exata após normalização
    2. O nome normalizado do app começa com a chave normalizada
       (fallback para quando a planilha usa nome curto)
    """
    app_norm = _normalizar(nome_setor_app)
    for meta in metas:
        chave_norm = _normalizar(meta["setor"])

        # Regra 1: igualdade exata
        if app_norm == chave_norm:
            return meta

        # Regra 2: prefixo (chave mais curta que o nome do app)
        if app_norm.startswith(chave_norm) and len(app_norm) > len(chave_norm):
            if app_norm[len(chave_norm)] in (" ", "/", "-"):
                return meta

    return None


# ---------------------------------------------------------------------------
# Meta diária (ritmo dentro do ciclo)
# ---------------------------------------------------------------------------

# Indicadores do aviso diário no Slack — escolha da gerência (ago/2026):
# Receita e Clientes Ativos. Os demais (multimarca, cabelo, make, RPA, %)
# ficam apenas na meta do ciclo. Para incluir outro, basta acrescentar a linha
# aqui — os acumulativos em clientes usam a meta Qtd da planilha, ex.:
#   ("clientes_cabelos", "meta_cabelos_qtd", "IAF Cabelo (clientes)", "int")
#   (chave_real, chave_meta, label, tipo)
# No bloco "Hoje", clientes = revendedoras com venda captada no dia.
INDICADORES_DIARIOS = [
    ("receita", "meta_receita", "Receita", "moeda"),
    ("clientes_ativos", "meta_ativo", "Clientes Ativos", "int"),
]


def _num(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def calcular_meta_diaria(dados: Dict[str, Any], posicao: Dict[str, Any]) -> Dict[str, Any]:
    """
    Quebra a meta do ciclo em ritmo diário para os indicadores acumulativos.

    Args:
        dados:   mesmo dict enviado ao Slack (receita, meta_receita, ...).
        posicao: saída de calendario_ciclos.posicao_ciclo.

    Para cada indicador com meta > 0:
        meta_dia        -> meta / dias úteis do ciclo (ritmo base)
        esperado        -> meta_dia * dia_atual (onde deveria estar hoje)
        falta           -> max(0, meta - real)
        necessario_dia  -> falta / dias_restantes (None se não restam dias)
        pct_meta        -> real / meta * 100
        pct_ritmo       -> real / esperado * 100 (None no dia 0)
        status          -> 'batida' | 'no_ritmo' | 'atrasado'
    status_geral: 'batida' (todas), 'atrasado' (alguma), 'no_ritmo', 'sem_meta'.
    """
    total = int(posicao.get("dias_uteis") or 0)
    dia_atual = int(posicao.get("dia_atual") or 0)
    restantes = int(posicao.get("dias_restantes") or 0)

    itens = []
    for chave_real, chave_meta, label, tipo in INDICADORES_DIARIOS:
        meta = _num(dados.get(chave_meta))
        if meta <= 0 or total <= 0:
            continue
        real = _num(dados.get(chave_real))
        meta_dia = meta / total
        esperado = meta_dia * dia_atual
        falta = max(0.0, meta - real)
        necessario_dia = (falta / restantes) if restantes > 0 else None
        pct_meta = real / meta * 100
        pct_ritmo = (real / esperado * 100) if esperado > 0 else None

        if real >= meta:
            status = "batida"
        elif esperado <= 0 or real >= esperado:
            status = "no_ritmo"
        else:
            status = "atrasado"

        itens.append({
            "chave": chave_real,
            "label": label,
            "tipo": tipo,
            "real": real,
            "meta": meta,
            "meta_dia": meta_dia,
            "esperado": esperado,
            "gap": real - esperado,
            "falta": falta,
            "necessario_dia": necessario_dia,
            "pct_meta": pct_meta,
            "pct_ritmo": pct_ritmo,
            "status": status,
        })

    if not itens:
        status_geral = "sem_meta"
    elif all(i["status"] == "batida" for i in itens):
        status_geral = "batida"
    elif any(i["status"] == "atrasado" for i in itens):
        status_geral = "atrasado"
    else:
        status_geral = "no_ritmo"

    return {"itens": itens, "status_geral": status_geral, "posicao": posicao}


# Tolerância (dias corridos) entre o início do ciclo e a primeira data da
# planilha para ela ainda contar como "acumulado do ciclo".
TOLERANCIA_ACUMULADO_DIAS = 2


def acumulado_valido(planilha: Optional[Dict[str, Any]], posicao: Dict[str, Any]) -> bool:
    """
    A planilha cobre o ciclo desde o início (→ o realizado é o acumulado)?

    planilha: saída de venda.obter_periodo_datas (data_min/data_max/n_dias).
    Sem informação de data assume-se acumulado (comportamento anterior).
    Uma planilha "só do dia" (data_min bem depois do início) NÃO serve para o
    ritmo acumulado — só para o resultado do dia.
    """
    if not planilha or not planilha.get("tem_data", True) or not planilha.get("data_min"):
        return True
    try:
        from datetime import date, timedelta
        data_min = date.fromisoformat(str(planilha["data_min"])[:10])
        inicio = date.fromisoformat(str(posicao.get("inicio", ""))[:10])
    except (ValueError, TypeError):
        return True
    return data_min <= inicio + timedelta(days=TOLERANCIA_ACUMULADO_DIAS)


def calcular_meta_do_dia(dados: Dict[str, Any], posicao: Dict[str, Any]) -> Dict[str, Any]:
    """
    Resultado DO DIA (recorte da planilha por DataCaptacao) contra a meta do
    dia (meta do ciclo ÷ dias úteis).

    dados["hoje"]: {data, receita, clientes_ativos, clientes_multimarcas,
                    clientes_cabelos, clientes_make} — None se não houver recorte.
    Para cada indicador com meta > 0:
        real_dia, meta_dia, pct (real_dia / meta_dia), status 'batida' | 'abaixo'
    status_geral: 'batida' (todas) | 'abaixo' | 'sem_meta' | 'sem_recorte'.
    """
    hoje = dados.get("hoje") or None
    total = int(posicao.get("dias_uteis") or 0)
    if not hoje or total <= 0:
        return {"itens": [], "status_geral": "sem_recorte", "data": None}

    itens = []
    for chave_real, chave_meta, label, tipo in INDICADORES_DIARIOS:
        meta = _num(dados.get(chave_meta))
        if meta <= 0:
            continue
        real_dia = _num(hoje.get(chave_real))
        meta_dia = meta / total
        pct = real_dia / meta_dia * 100 if meta_dia > 0 else 0.0
        itens.append({
            "chave": chave_real,
            "label": label,
            "tipo": tipo,
            "real_dia": real_dia,
            "meta_dia": meta_dia,
            "pct": pct,
            "status": "batida" if real_dia >= meta_dia else "abaixo",
        })

    if not itens:
        status_geral = "sem_meta"
    elif all(i["status"] == "batida" for i in itens):
        status_geral = "batida"
    else:
        status_geral = "abaixo"
    return {"itens": itens, "status_geral": status_geral, "data": hoje.get("data")}
